"""
El lector de Excel: las seis trampas que el CSV no tiene.

🔴 LOS ARCHIVOS SON REALES, generados con openpyxl dentro del test y leídos como bytes — no hay
un fake del lector en ningún lado. Es la única forma de probar esto: el modo de falla que se
cubre acá NO es lógico sino de FORMATO (qué tipo devuelve openpyxl para cada celda), así que un
doble que devolviera dicts de strings prefabricados haría pasar todos los tests con el lector
borrado. El archivo binario es el fake, y es el único que puede desmentir.

⚠️ ¿QUÉ TENDRÍA QUE SER DISTINTO EN EL ARCHIVO PARA QUE ESTOS TESTS PUEDAN FALLAR?

  1. 🔴 CADA TRAMPA ESTÁ EN EL ARCHIVO, NO SOLO EN EL ASSERT. La planilla base trae una celda
     vacía, un número, una fecha, un header con espacio final, DOS hojas y una fila fantasma al
     final. Si el archivo tuviera solo strings limpios en una hoja, un lector que no maneje
     ninguna de las seis pasaría igual — que es exactamente el caso #1 de la regla del repo.
  2. 🔴 LOS VALORES SON DISTINGUIBLES DE SU MODO DE FALLA. La celda vacía se compara contra
     `""` y también contra `"None"` (que es lo que saldría de `str(None)`); el número contra
     `"1"` y también contra `"1.0"`; la fecha contra `d/m/Y` y no contra el `repr` del
     datetime. Sin ese segundo assert, "lo maneja" y "lo rompe de otra forma" darían lo mismo.
  3. 🔴 LA SEGUNDA HOJA TIENE CONTENIDO PROPIO Y DISTINTO. Con una hoja vacía, "lee la primera"
     y "lee la última" serían indistinguibles.
  4. La fila fantasma va DESPUÉS de una fila con datos, así que descartarla no puede confundirse
     con "cortar al primer vacío".
"""
import os

_TEST_ENV: dict[str, str] = {
    "SUPABASE_URL": "https://test-project.supabase.co",
    "SUPABASE_ANON_KEY": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.test.anon",
    "SUPABASE_SERVICE_KEY": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.test.service",
    "JWT_SECRET": "test-secret-for-unit-tests-only-minimum-32-chars!!",
    "ANTHROPIC_API_KEY": "sk-ant-test",
    "RESEND_API_KEY": "re_test",
}
for _k, _v in _TEST_ENV.items():
    os.environ.setdefault(_k, _v)

import io  # noqa: E402
from datetime import date  # noqa: E402

import pytest  # noqa: E402
from openpyxl import Workbook  # noqa: E402

from services import _import_excel as lector  # noqa: E402


def _xlsx(hojas: dict) -> bytes:
    """Arma un .xlsx real. `hojas` = {nombre: [fila, fila, ...]}, la primera es la primera."""
    wb = Workbook()
    primera = True
    for nombre, filas in hojas.items():
        ws = wb.active if primera else wb.create_sheet(nombre)
        if primera:
            ws.title = nombre
            primera = False
        for f in filas:
            ws.append(f)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# 🔴 Punto 1 del encabezado: las seis trampas viven en ESTE archivo, no en los asserts.
def _archivo_con_las_trampas() -> bytes:
    return _xlsx({
        "Datos": [
            ["Titulo ", "Cantidad", "Fecha", "Nota"],       # trampa 4: espacio final
            ["Migrar nómina", 12345678, date(2026, 6, 30), None],  # trampas 1, 2 y 3
            ["Otra tarea", 2.5, None, "con nota"],
            [None, None, None, None],                       # trampa 6: fila fantasma
        ],
        "Segunda": [["NO", "DEBE", "LEERSE"], ["basura", 1, 2]],   # trampa 5
    })


@pytest.fixture
def datos() -> bytes:
    return _archivo_con_las_trampas()


# ── 0. El guardián del archivo ────────────────────────────────────────────────

def test_el_archivo_de_prueba_trae_las_seis_trampas(datos) -> None:
    """Sin esto, todo lo de abajo puede volverse decorativo si alguien 'limpia' la planilla."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(datos), read_only=True, data_only=True)
    assert wb.sheetnames == ["Datos", "Segunda"]            # dos hojas, con contenido distinto
    crudas = list(wb.worksheets[0].iter_rows(values_only=True))
    assert crudas[0][0] == "Titulo "                        # header con espacio
    assert crudas[1][3] is None                             # celda vacía
    assert isinstance(crudas[1][1], int)                    # número
    assert crudas[1][2] is not None                         # fecha
    assert all(v is None for v in crudas[3])                # fila fantasma


# ── 1. Las seis trampas ───────────────────────────────────────────────────────

class TestLasTrampasDeExcel:

    def test_1_una_celda_vacia_sale_como_string_vacio_y_NUNCA_como_None(self, datos) -> None:
        """🔴 `str(None)` da "None", que entraría a la base como ese texto literal."""
        headers, crudas = lector.abrir(datos)
        primera = dict(lector.filas(headers, crudas))[2]

        assert primera["nota"] == ""
        assert primera["nota"] != "None"

    def test_2_un_numero_entero_no_arrastra_el_punto_cero(self, datos) -> None:
        """🔴 El modo de falla real: un DNI `12345678` que sale `"12345678.0"` no matchea NUNCA
        en un dedup, y el import crea duplicados sin un solo error."""
        headers, crudas = lector.abrir(datos)
        primera = dict(lector.filas(headers, crudas))[2]

        assert primera["cantidad"] == "12345678"
        assert primera["cantidad"] != "12345678.0"

    def test_2_bis_un_float_integral_de_OTRO_productor_tampoco_arrastra_el_punto_cero(self) -> None:
        """🔴 ESTE TEST EXISTE POR UN AGUJERO QUE ENCONTRÓ LA MUTACIÓN, y hay que decir por qué.

        El test de arriba pasa el archivo real por el lector, pero **openpyxl normaliza a `int`
        los floats integrales que él mismo escribe**: con una planilla generada acá, la rama
        `isinstance(v, float) and v.is_integer()` NUNCA se ejecuta, y borrarla dejaba todo en
        verde. O sea que el test end-to-end no podía desmentir la trampa 2.

        El caso real llega de OTROS productores —un export de otro sistema, una conversión de
        CSV— que sí escriben `<v>12345678.0</v>`. Como no se puede fabricar ese archivo con
        openpyxl, la normalización se prueba DIRECTO sobre `_valor`, que es el único punto donde
        la rama es alcanzable."""
        assert lector._valor(12345678.0) == "12345678"
        assert lector._valor(12345678.0) != "12345678.0"
        assert lector._valor(2.5) == "2.5"          # contrapeso: no rompe los decimales

    def test_2b_un_decimal_de_verdad_conserva_sus_decimales(self, datos) -> None:
        """Contrapeso: un lector que hiciera `int()` a todo rompería los importes."""
        headers, crudas = lector.abrir(datos)
        segunda = dict(lector.filas(headers, crudas))[3]

        assert segunda["cantidad"] == "2.5"

    def test_3_una_fecha_sale_formateada_y_no_como_datetime(self, datos) -> None:
        """Sale en `d/m/Y`, que es el formato que ya leen los parsers del repo."""
        headers, crudas = lector.abrir(datos)
        primera = dict(lector.filas(headers, crudas))[2]

        assert primera["fecha"] == "30/06/2026"
        assert "datetime" not in primera["fecha"] and "00:00" not in primera["fecha"]

    def test_4_un_header_con_espacio_final_se_encuentra_igual(self, datos) -> None:
        """🔴 `"Titulo "` es indistinguible de `"Titulo"` en pantalla: sin normalizar, el error
        diría "falta la columna Titulo" con la columna a la vista."""
        headers, crudas = lector.abrir(datos)

        assert lector.faltantes(headers, ["Titulo"]) == []
        assert dict(lector.filas(headers, crudas))[2]["titulo"] == "Migrar nómina"

    def test_5_se_lee_SOLO_la_primera_hoja(self, datos) -> None:
        """🔴 La primera y no la activa: la activa es la pestaña donde quedó el cursor al
        guardar, o sea que el archivo se leería distinto según cómo lo cerró quien lo mandó."""
        headers, crudas = lector.abrir(datos)
        leidas = dict(lector.filas(headers, crudas))

        assert "basura" not in str(leidas)
        assert lector.hojas(datos) == ["Datos", "Segunda"]   # las ve, pero lee una

    def test_6_las_filas_fantasma_del_final_se_descartan(self, datos) -> None:
        """Sin esto, el reporte diría "3 filas con error" sobre un archivo de 2 datos."""
        headers, crudas = lector.abrir(datos)

        assert len(dict(lector.filas(headers, crudas))) == 2


# ── 2. La numeración y las cabeceras ──────────────────────────────────────────

class TestNumeracionYCabeceras:

    def test_la_primera_fila_de_datos_es_la_2(self, datos) -> None:
        """Es el número que ve quien abre el Excel: la 1 es el encabezado."""
        headers, crudas = lector.abrir(datos)

        assert [n for n, _ in lector.filas(headers, crudas)] == [2, 3]

    def test_una_fila_vacia_en_el_MEDIO_no_corre_la_numeracion(self) -> None:
        """🔴 Si descartar una fila corriera los números, el reporte mandaría al usuario a la
        fila equivocada — y el reporte existe justamente para que vaya a la correcta."""
        datos = _xlsx({"D": [["Titulo"], ["uno"], [None], ["tres"]]})
        headers, crudas = lector.abrir(datos)

        assert [n for n, _ in lector.filas(headers, crudas)] == [2, 4]

    def test_faltantes_devuelve_el_nombre_ORIGINAL_de_la_columna(self, datos) -> None:
        """El mensaje lo lee alguien con la planilla abierta: "nota final" no se encuentra,
        "NOTA FINAL" sí."""
        headers, _ = lector.abrir(datos)

        assert lector.faltantes(headers, ["NOTA FINAL", "Titulo"]) == ["NOTA FINAL"]

    def test_faltantes_usa_la_MISMA_normalizacion_que_el_CSV(self, datos) -> None:
        """Las dos rutas tienen que aceptar y rechazar exactamente los mismos encabezados."""
        from services._import_csv import normalizar_header

        headers, _ = lector.abrir(datos)
        assert normalizar_header("  TITULO  ") == normalizar_header("Titulo ")
        assert lector.faltantes(headers, ["  TITULO  "]) == []


# ── 3. Archivos que no se pueden leer ─────────────────────────────────────────

class TestArchivosRotos:

    def test_un_archivo_que_no_es_excel_da_un_mensaje_legible(self) -> None:
        """No un traceback de openpyxl: el que lo lee es alguien de RRHH."""
        with pytest.raises(ValueError) as exc:
            lector.abrir(b"esto,es,un,csv\n1,2,3,4")

        assert ".xlsx" in str(exc.value)

    def test_un_excel_sin_ninguna_fila_no_rompe(self) -> None:
        headers, crudas = lector.abrir(_xlsx({"Vacía": []}))

        assert headers == [] and list(lector.filas(headers, crudas)) == []

    def test_hojas_de_un_archivo_roto_devuelve_lista_vacia(self) -> None:
        """Se usa solo para informar cuál se leyó: no puede tumbar el import."""
        assert lector.hojas(b"no soy un excel") == []
