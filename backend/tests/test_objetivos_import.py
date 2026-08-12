"""
Import de objetivos por Excel: preview → confirmar.

⚠️ ¿QUÉ TENDRÍA QUE SER DISTINTO EN LOS FAKES PARA QUE ESTOS TESTS PUEDAN FALLAR?

  1. 🔴 LOS ARCHIVOS SON .xlsx REALES, armados con openpyxl acá adentro. El lector no está
     falseado: si lo estuviera, "el import lee el Excel" y "el import lee un dict que le
     pasamos" serían indistinguibles.
  2. 🔴 `users` FALSA TIENE UN ACTIVO Y UN INACTIVO, y hay un tercer nombre que no existe. Con
     un solo usuario activo, las dos ramas de rechazo —"no existe" y "está inactivo"— darían el
     mismo resultado y borrar cualquiera de las dos quedaría en verde.
  3. 🔴 EL ARCHIVO DEL LOTE TRAE 3 FILAS. Con una sola, "un evento por lote" y "un evento por
     fila" son EL MISMO número: el test no podría distinguirlos, que es justo lo que la regla
     del repo prohíbe.
  4. 🔴 EL SERVICE DE OBJETIVOS FALSO REGISTRA CADA `create` Y FALLA SELECTIVAMENTE. Registrar
     permite contar altas; fallar en una sola fila permite verificar que el lote NO aborta y que
     la fila fallida se reporta — con un fake que siempre acepta, esa rama no se ejecuta nunca.
  5. El auditor falso registra los eventos: se afirma CUÁNTOS se emitieron y con qué payload, no
     solo que la llamada no explotó.
"""
import os

_TEST_ENV: dict[str, str] = {
    "SUPABASE_URL": "https://test-project.supabase.co",
    "SUPABASE_ANON_KEY": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.test.anon",
    "SUPABASE_SERVICE_KEY": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.test.service",
    "JWT_SECRET": "test-secret-for-unit-tests-only-minimum-32-chars!!",
    "ANTHROPIC_API_KEY": "sk-ant-test",
    "RESEND_API_KEY": "re_test",
}
for _k, _v in _TEST_ENV.items():
    os.environ.setdefault(_k, _v)

import io  # noqa: E402
from datetime import date, datetime  # noqa: E402
from types import SimpleNamespace  # noqa: E402
from uuid import uuid4  # noqa: E402

import pytest  # noqa: E402
from openpyxl import Workbook  # noqa: E402

import services.objetivos_import_preview as prev_mod  # noqa: E402
from schemas.importacion_objetivos import (  # noqa: E402
    ImportacionObjetivosConfirmarRequest,
)
from schemas.objetivo import ObjetivoResponse  # noqa: E402
from services.objetivos_import_service import ObjetivosImportService  # noqa: E402
from utils.errors import AppError  # noqa: E402

EMPRESA = str(uuid4())
ANA = str(uuid4())
BETO = str(uuid4())

# 🔴 Punto 2: un ACTIVO, un INACTIVO y un tercero que no existe.
_USERS = [
    {"id": ANA, "nombre": "Ana", "apellido": "Gómez", "email": "ana@karstec.com",
     "username": "agomez", "activo": True},
    {"id": BETO, "nombre": "Beto", "apellido": "Pérez", "email": "beto@karstec.com",
     "username": "bperez", "activo": True},
    {"id": str(uuid4()), "nombre": "Caro", "apellido": "Díaz", "email": "caro@karstec.com",
     "username": "cdiaz", "activo": False},
]


def _xlsx(filas) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Objetivos"
    for f in filas:
        ws.append(f)
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


_HEADERS = ["Titulo", "Responsable", "Prioridad", "Fecha entrega", "Descripcion", "Responsables"]


def _archivo_de_3_filas() -> bytes:
    """🔴 Punto 3: TRES filas. Con una, "un evento por lote" y "uno por fila" son el mismo 1."""
    return _xlsx([
        _HEADERS,
        ["Migrar nómina", "ana@karstec.com", "alta", date(2026, 6, 30), "Q2", "beto@karstec.com"],
        ["Auditar licencias", "agomez", "media", None, None, None],
        ["Revisar convenio", "Beto Pérez", "baja", date(2026, 9, 1), None, None],
    ])


class _UsersFake:
    """La tabla `users` como la consulta el preview: select→eq(activo)→order→execute.

    ⚠️ Desde la sesión 0.9 el preview no consulta `users` directo: usa
    `UsuarioRepo.listar_activos()`, que ya existía para el listado y su export. Por eso se
    faltea `usuario_repo` y el fake suma `order` — ese método ordena por apellido. El orden no
    cambia el resultado de este test: el preview arma un índice por email/username/nombre, y la
    clave no depende de la posición.
    """

    def table(self, nombre: str):
        assert nombre == "users", f"el preview consultó {nombre!r}, no users"
        return self

    def select(self, *a, **k):
        return self

    def order(self, *a, **k):
        return self

    def eq(self, col, val):
        self._activo = val
        return self

    def execute(self):
        # Honra el filtro: con `activo=True` el inactivo NO sale. Si lo ignorara, la rama de
        # "responsable inactivo" sería inalcanzable y su test pasaría con el filtro borrado.
        return SimpleNamespace(data=[u for u in _USERS if u["activo"] == self._activo])


class _ObjetivosFake:
    """🔴 Punto 4: registra los `create` y puede fallar en una fila concreta."""

    def __init__(self, falla_en: str = "") -> None:
        self.creados: list = []
        self._falla_en = falla_en

    def create(self, data, created_by):
        if self._falla_en and data.titulo == self._falla_en:
            raise AppError("El responsable no está activo", "RESPONSABLE_NO_ACTIVO", 422)
        self.creados.append(data)
        return ObjetivoResponse(
            id=str(uuid4()), empresa_id=str(data.empresa_id), responsable_id=str(data.responsable_id),
            titulo=data.titulo, prioridad=data.prioridad, estado="por_hacer",
            fecha_entrega=data.fecha_entrega, created_at=datetime(2026, 1, 1),
            updated_at=datetime(2026, 1, 1),
        )


class _AuditFake:
    def __init__(self) -> None:
        self.eventos: list = []

    def registrar(self, **kw):
        self.eventos.append(kw)


@pytest.fixture
def users(monkeypatch):
    import repositories.usuario_repo as usuario_repo_mod
    monkeypatch.setattr(usuario_repo_mod, "supabase_admin", _UsersFake())


def _svc(falla_en: str = ""):
    objetivos, audit = _ObjetivosFake(falla_en), _AuditFake()
    return ObjetivosImportService(objetivos=objetivos, audit=audit), objetivos, audit


# ── 0. Los guardianes del fake ────────────────────────────────────────────────

def test_el_fake_de_users_tiene_un_activo_y_un_inactivo() -> None:
    assert [u["activo"] for u in _USERS] == [True, True, False]


def test_el_archivo_del_lote_tiene_tres_filas(users) -> None:
    """Sin 3 filas, "un evento por lote" no se puede distinguir de "uno por fila"."""
    svc, _, _ = _svc()
    assert len(svc.preview(_archivo_de_3_filas()).filas_validas) == 3


# ── 1. Preview: la fila válida ────────────────────────────────────────────────

class TestPreviewFilaValida:

    def test_una_fila_valida_resuelve_el_responsable(self, users) -> None:
        svc, _, _ = _svc()

        f = svc.preview(_archivo_de_3_filas()).filas_validas[0]

        assert f.titulo == "Migrar nómina"
        # `str(...)`: `responsable_id` es UUID en el schema desde la sesión 0.6, y ANA es el str
        # que el fake de `users` devuelve. Comparar el valor —no el tipo— es lo que este test mira.
        assert str(f.responsable_id) == ANA and f.responsable_nombre == "Ana Gómez"

    def test_el_responsable_se_puede_escribir_como_email_username_o_nombre(self, users) -> None:
        """Las tres claves porque el archivo lo escribe una persona: no hay forma de saber cuál
        va a usar. Las tres filas del archivo usan una distinta."""
        svc, _, _ = _svc()

        ids = [str(f.responsable_id) for f in svc.preview(_archivo_de_3_filas()).filas_validas]

        assert ids == [ANA, ANA, BETO]   # email, username, "nombre apellido"

    def test_la_fecha_llega_convertida_a_iso(self, users) -> None:
        svc, _, _ = _svc()

        assert svc.preview(_archivo_de_3_filas()).filas_validas[0].fecha_entrega == "2026-06-30"

    def test_los_responsables_ADICIONALES_se_resuelven_y_no_repiten_al_dueño(self, users) -> None:
        """🔴 La puente (migración 096). El dueño entra por su propio camino: duplicarlo acá
        chocaría contra la PK compuesta de `objetivo_responsables`."""
        svc, _, _ = _svc()

        f = svc.preview(_archivo_de_3_filas()).filas_validas[0]

        assert f.responsables_ids == [BETO]
        assert ANA not in f.responsables_ids

    def test_una_fila_sin_fecha_se_carga_igual(self, users) -> None:
        svc, _, _ = _svc()

        f = svc.preview(_archivo_de_3_filas()).filas_validas[1]

        assert f.fecha_entrega is None and f.faltantes == []


# ── 2. Preview: lo que NO se carga ────────────────────────────────────────────

class TestPreviewRechazos:

    def test_un_responsable_INEXISTENTE_se_reporta_y_no_se_carga(self, users) -> None:
        """🔴 Nunca se carga con responsable nulo: la columna es NOT NULL y el insert daría un
        500 en vez de un renglón accionable en el reporte."""
        svc, _, _ = _svc()
        datos = _xlsx([_HEADERS, ["Tarea", "nadie@karstec.com", "alta", None, None, None]])

        res = svc.preview(datos)

        assert res.filas_validas == []
        assert len(res.errores) == 1
        assert "nadie@karstec.com" in res.errores[0].motivo
        assert res.errores[0].fila == 2 and res.errores[0].identificador == "Tarea"

    def test_un_responsable_INACTIVO_se_reporta_igual_que_uno_inexistente(self, users) -> None:
        """Punto 2 del encabezado: sin el usuario inactivo en el fake, esta rama es inalcanzable."""
        svc, _, _ = _svc()
        datos = _xlsx([_HEADERS, ["Tarea", "caro@karstec.com", "alta", None, None, None]])

        res = svc.preview(datos)

        assert res.filas_validas == [] and len(res.errores) == 1

    def test_un_acompañante_inexistente_tumba_la_fila_entera(self, users) -> None:
        """Contrapeso: cargar el objetivo sin ese responsable sería perder un dato en silencio."""
        svc, _, _ = _svc()
        datos = _xlsx([_HEADERS, ["Tarea", "ana@karstec.com", "alta", None, None, "fantasma@x.com"]])

        res = svc.preview(datos)

        assert res.filas_validas == [] and "fantasma@x.com" in res.errores[0].motivo

    def test_una_fila_sin_titulo_se_reporta(self, users) -> None:
        svc, _, _ = _svc()
        datos = _xlsx([_HEADERS, [None, "ana@karstec.com", "alta", None, None, None]])

        res = svc.preview(datos)

        assert res.filas_validas == [] and "título" in res.errores[0].motivo.lower()

    def test_las_filas_buenas_del_mismo_archivo_SI_se_cargan(self, users) -> None:
        """🔴 El lote no aborta por una fila: es la mitad que hace útil el reporte."""
        svc, _, _ = _svc()
        datos = _xlsx([
            _HEADERS,
            ["Buena", "ana@karstec.com", "alta", None, None, None],
            ["Mala", "nadie@x.com", "alta", None, None, None],
            ["Otra buena", "bperez", "baja", None, None, None],
        ])

        res = svc.preview(datos)

        assert [f.titulo for f in res.filas_validas] == ["Buena", "Otra buena"]
        assert [e.fila for e in res.errores] == [3]


# ── 3. Preview: el archivo entero ─────────────────────────────────────────────

class TestArchivoEntero:

    def test_si_falta_una_columna_requerida_se_rechaza_TODO(self, users) -> None:
        """🔴 Ni una fila se procesa: procesar la mitad de un archivo mal armado deja media
        carga hecha que después hay que deshacer a mano."""
        svc, objetivos, _ = _svc()
        datos = _xlsx([["Titulo", "Prioridad"], ["Tarea", "alta"]])

        with pytest.raises(AppError) as exc:
            svc.preview(datos)

        assert exc.value.code == "COLUMNAS_FALTANTES"
        assert "Responsable" in exc.value.message
        assert objetivos.creados == []

    def test_un_archivo_con_columna_de_PADRE_se_rechaza_entero(self, users) -> None:
        """🔴 La jerarquía no se soporta y NO se ignora en silencio: aceptar el archivo y
        descartar la columna haría que el usuario crea que cargó una jerarquía y cargue una
        lista plana, sin enterarse nunca."""
        svc, _, _ = _svc()
        datos = _xlsx([[*_HEADERS, "Objetivo padre"],
                       ["Hija", "ana@karstec.com", "alta", None, None, None, "Madre"]])

        with pytest.raises(AppError) as exc:
            svc.preview(datos)

        assert exc.value.code == "JERARQUIA_NO_SOPORTADA"
        assert "pantalla" in exc.value.message

    def test_un_archivo_que_no_es_excel_da_422_y_no_500(self, users) -> None:
        svc, _, _ = _svc()

        with pytest.raises(AppError) as exc:
            svc.preview(b"titulo;responsable\nTarea;ana")

        assert exc.value.code == "EXCEL_ILEGIBLE" and exc.value.status_code == 422

    def test_informa_que_hoja_leyo_cuando_hay_varias(self, users) -> None:
        """Para que no sorprenda: el usuario mandó tres pestañas y se leyó una."""
        wb = Workbook(); wb.active.title = "Primera"
        for f in [_HEADERS, ["T", "ana@karstec.com", "alta", None, None, None]]:
            wb.active.append(f)
        wb.create_sheet("Otra").append(["x"])
        buf = io.BytesIO(); wb.save(buf)
        svc, _, _ = _svc()

        res = svc.preview(buf.getvalue())

        assert res.hoja_leida == "Primera" and res.total_hojas == 2


# ── 4. 🔴 Confirmar: UN evento por lote, no uno por fila ──────────────────────

class TestConfirmar:

    def _body(self, svc, users_fixture=None):
        return ImportacionObjetivosConfirmarRequest(
            empresa_id=EMPRESA, filas=svc.preview(_archivo_de_3_filas()).filas_validas)

    def test_carga_las_tres_filas(self, users) -> None:
        svc, objetivos, _ = _svc()

        res = svc.confirmar(self._body(svc), "operador-1")

        assert res.importados == 3 and len(objetivos.creados) == 3

    def test_UN_evento_de_auditoria_para_TRES_filas(self, users) -> None:
        """🔴 La regla del repo: al auditar una importación, UN evento por lote. Con 3 filas y
        3 eventos, el log de auditoría se vuelve ilegible el día que el archivo traiga 200."""
        svc, _, audit = _svc()

        svc.confirmar(self._body(svc), "operador-1")

        assert len(audit.eventos) == 1

    def test_el_evento_lleva_la_empresa_del_BODY_y_un_id_de_evento(self, users) -> None:
        """La empresa del body y no del header: confirmar es una ACCIÓN. Y `registro_id` es un
        uuid4 de EVENTO, no de recurso: el import no persiste un lote con id propio."""
        svc, _, audit = _svc()

        svc.confirmar(self._body(svc), "operador-1")
        ev = audit.eventos[0]

        assert ev["empresa_id"] == EMPRESA
        assert ev["entidad"] == "objetivo" and ev["evento"] == "importacion_objetivos"
        assert ev["registro_id"] and ev["registro_id"] not in [str(EMPRESA)]

    def test_el_evento_nombra_los_ids_creados(self, users) -> None:
        """Un alta es una fotografía y el id alcanza para reconstruirla; sin los ids, "se
        importaron 3 objetivos" no permitiría saber cuáles."""
        svc, _, audit = _svc()

        svc.confirmar(self._body(svc), "operador-1")

        assert len(audit.eventos[0]["datos_nuevos"]["ids_creados"]) == 3

    def test_una_fila_que_falla_NO_aborta_el_lote_y_se_reporta(self, users) -> None:
        """🔴 Punto 4 del encabezado: el fake falla en UNA fila concreta. Con un fake que
        siempre acepta, esta rama no se ejecuta nunca."""
        svc, objetivos, audit = _svc(falla_en="Auditar licencias")

        res = svc.confirmar(self._body(svc), "operador-1")

        assert res.importados == 2 and len(objetivos.creados) == 2
        assert [e.identificador for e in res.errores] == ["Auditar licencias"]
        assert len(audit.eventos) == 1                     # sigue siendo UNO

    def test_el_evento_se_emite_TAMBIEN_con_el_lote_vacio(self, users) -> None:
        """Es el único registro de que alguien importó: sin él, un import sin filas válidas no
        deja una sola línea en `auditoria`."""
        svc, _, audit = _svc()

        res = svc.confirmar(ImportacionObjetivosConfirmarRequest(empresa_id=EMPRESA, filas=[]))

        assert res.importados == 0 and len(audit.eventos) == 1

    def test_el_confirmar_pasa_los_responsables_adicionales(self, users) -> None:
        """La puente se escribe por el alta, no por el import: `ObjetivoService.create` la
        resuelve. Lo que se verifica acá es que el import no los pierda en el camino."""
        svc, objetivos, _ = _svc()

        svc.confirmar(self._body(svc), "operador-1")

        assert [str(u) for u in objetivos.creados[0].responsables] == [BETO]

    def test_ningun_objetivo_se_crea_como_HIJO(self, users) -> None:
        """La jerarquía no se soporta: todo nace raíz. Si algún día el import la arme, este
        test rojea y obliga a decidirlo a propósito."""
        svc, objetivos, _ = _svc()

        svc.confirmar(self._body(svc), "operador-1")

        assert all(o.parent_id is None for o in objetivos.creados)
