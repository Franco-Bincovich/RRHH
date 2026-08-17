"""
Las tres columnas nuevas del import (Tipo · Periodicidad · Areas involucradas), el desplegable de
áreas ya usadas, y el mensaje del duplicado.

Vive aparte de `test_objetivos_import.py` (409 líneas) porque cubre una superficie nueva y porque
dos de sus bloques necesitan fakes que allá no existen: uno que devuelva ARRAYS para el pool de
áreas, y un repo que rebote con un 23505 real.

## ⚠️ ¿QUÉ TENDRÍA QUE SER DISTINTO EN LOS FAKES PARA QUE ESTOS TESTS PUEDAN FALLAR?

  1. 🔴 **LOS .xlsx SON REALES**, armados con openpyxl acá adentro, y hay UNO SIN las columnas
     nuevas y otro CON ellas. Con un solo archivo, "el import lee la columna" y "el import usa el
     default" serían indistinguibles: los dos darían el mismo resultado para el archivo viejo.
  2. 🔴 **EL ARCHIVO NUEVO TRAE UN `Tipo` VÁLIDO Y UNO INVÁLIDO.** Sin el válido, un
     `parsear_fila` que devolviera SIEMPRE el default pasaría el test del inválido — que es el
     caso #1 de la regla del repo con otra cara.
  3. 🔴 **HAY UN ÁREA CON COMA ADENTRO Y OTRA SIN**, y las celdas usan los dos separadores. Sin
     ese par, la regla de prioridad del `;` y una regla que partiera por los dos a la vez dan el
     mismo resultado.
  4. 🔴 **EL CATÁLOGO DEL POOL TIENE DOS FILAS QUE COMPARTEN UN ÁREA** y una con el array vacío.
     Sin la repetición no se puede ver que deduplica; sin la vacía, que no aporta un `''`.
  5. 🔴 **EL REPO QUE REBOTA EXPONE `.code`** como la `APIError` de postgrest, y el `create` que
     lo consume es el REAL — si fuera un doble, la traducción a 409 no correría y el test estaría
     verificando el fake.

## 🔴 Y LA LECCIÓN DE LA SESIÓN ANTERIOR: SOSPECHAR DE LA PROYECCIÓN

En la sesión 1, `test_NO_matchea_un_PREFIJO` no rojeó ante su mutación porque afirmaba sobre los
títulos RAÍZ y el hijo quedaba anidado: el `not in` se cumplía por la forma del árbol, no por el
filtro. Acá la trampa equivalente es afirmar sobre `len(filas_validas)` o sobre "no hubo error":
los dos se cumplen con los tres campos nuevos perdidos en el camino. **Todas las aserciones de
este archivo miran el VALOR de los tres campos, y las del confirmar los miran en el
`ObjetivoCreate` que llega al repo** — que es el último punto antes de la base.
"""
import os

_TEST_ENV: dict[str, str] = {
    "SUPABASE_URL": "https://test-project.supabase.co",
    "SUPABASE_ANON_KEY": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.test.anon",
    "SUPABASE_SERVICE_KEY": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.test.service",
    "JWT_SECRET": "test-secret-for-unit-tests-only-minimum-32-chars!!",
    "ANTHROPIC_API_KEY": "sk-ant-test",
    "RESEND_API_KEY": "re_test",
}
for _k, _v in _TEST_ENV.items():
    os.environ.setdefault(_k, _v)

import io  # noqa: E402
from datetime import date, datetime  # noqa: E402
from types import SimpleNamespace  # noqa: E402
from uuid import UUID, uuid4  # noqa: E402

import pytest  # noqa: E402
from openpyxl import Workbook  # noqa: E402

import services.objetivos_import_preview as prev_mod  # noqa: E402
from repositories.objetivo_areas_repo import ObjetivoAreasRepo  # noqa: E402
from schemas.importacion_objetivos import (  # noqa: E402
    ImportacionObjetivosConfirmarRequest,
)
from schemas.objetivo import ObjetivoResponse  # noqa: E402
from services._objetivos_export import construir_filas_export  # noqa: E402
from services._objetivos_import_transforms import parsear_fila  # noqa: E402
from services._objetivos_import_valores import separar_areas  # noqa: E402
from services.objetivo_catalogos_service import ObjetivoCatalogosService  # noqa: E402
from services.objetivo_service import ObjetivoService  # noqa: E402
from services.objetivos_import_service import ObjetivosImportService  # noqa: E402
from tests._fake_supabase import FakeSupabase  # noqa: E402
from utils.errors import AppError  # noqa: E402

EMPRESA = str(uuid4())
ANA = str(uuid4())

_USERS = [{"id": ANA, "nombre": "Ana", "apellido": "Gómez", "email": "ana@karstec.com",
           "username": "agomez", "activo": True}]


def _xlsx(filas) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Objetivos"
    for f in filas:
        ws.append(f)
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


# 🔴 Punto 1: el archivo VIEJO, tal como existe hoy en la máquina de RRHH. Sus seis columnas son
# exactamente las que el import aceptaba antes de la migración 119.
_HEADERS_VIEJOS = ["Titulo", "Responsable", "Prioridad", "Fecha entrega", "Descripcion",
                   "Responsables"]


def _archivo_viejo() -> bytes:
    return _xlsx([
        _HEADERS_VIEJOS,
        ["Migrar nómina", "ana@karstec.com", "alta", date(2026, 6, 30), "Q2", None],
        ["Auditar licencias", "agomez", "media", None, None, None],
    ])


# 🔴 Puntos 2 y 3: un Tipo válido, uno inválido y uno vacío; áreas con `;`, con `,` y con una coma
# ADENTRO de un área. Cada fila ejercita una rama distinta y ninguna repite la de al lado.
_HEADERS_NUEVOS = _HEADERS_VIEJOS + ["Tipo", "Periodicidad", "Areas involucradas"]


def _archivo_nuevo() -> bytes:
    return _xlsx([
        _HEADERS_NUEVOS,
        # Tipo válido + área con coma adentro, separada con `;` (como la escribe el export)
        ["Plan anual", "ana@karstec.com", "alta", None, None, None,
         "anual", "", "Legales, Compliance; Sistemas"],
        # Tipo INVÁLIDO -> cae al default, la fila entra igual
        ["Cerrar el mes", "agomez", "media", None, None, None,
         "semestral", "mensual", "Sistemas"],
        # Tipo con mayúscula + separador coma (copiar y pegar) + sin áreas
        ["Revisar convenio", "ana@karstec.com", "baja", None, None, None,
         "Anual", "", "Sistemas, Legales"],
    ])


class _UsersFake:
    def table(self, nombre: str):
        assert nombre == "users", f"el preview consultó {nombre!r}, no users"
        return self

    def select(self, *a, **k):
        return self

    def order(self, *a, **k):
        return self

    def eq(self, col, val):
        self._activo = val
        return self

    def execute(self):
        return SimpleNamespace(data=[u for u in _USERS if u["activo"] == self._activo])


@pytest.fixture
def users(monkeypatch):
    import repositories.usuario_repo as repo_mod
    monkeypatch.setattr(repo_mod, "supabase_admin", _UsersFake())
    return prev_mod


# ── 0. Guardianes del fake ────────────────────────────────────────────────────

def test_el_archivo_viejo_NO_trae_las_tres_columnas() -> None:
    """Punto 1. Si las trajera, "usa el default" y "lee la columna" serían el mismo resultado."""
    for col in ("Tipo", "Periodicidad", "Areas involucradas"):
        assert col not in _HEADERS_VIEJOS


def test_el_archivo_nuevo_trae_un_Tipo_VALIDO_y_uno_INVALIDO() -> None:
    """Punto 2. Sin el válido, un parser que devolviera siempre el default pasaría igual."""
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(_archivo_nuevo())).active
    tipos = [ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)]
    assert "anual" in tipos and "semestral" in tipos


def test_el_archivo_nuevo_trae_un_area_con_COMA_adentro() -> None:
    """Punto 3. Sin ella, la regla del `;` y partir por los dos dan lo mismo."""
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(_archivo_nuevo())).active
    celdas = [ws.cell(row=r, column=9).value for r in range(2, ws.max_row + 1)]
    assert any("," in (c or "") and ";" in (c or "") for c in celdas)


# ── 1. 🔴 El archivo viejo sigue entrando igual ───────────────────────────────

class TestElArchivoViejoEntraCompleto:
    """🔴 EL TEST DE LA SESIÓN. Un archivo anterior a la migración 119 no puede empezar a
    rebotar: es literalmente el que RRHH tiene guardado en su máquina."""

    def test_las_dos_filas_entran_sin_errores(self, users) -> None:
        res = prev_mod.preview(_archivo_viejo())

        assert len(res.filas_validas) == 2
        assert res.errores == []

    def test_los_defaults_son_los_correctos(self, users) -> None:
        """⚠️ Se afirma sobre el VALOR de los tres campos, no sobre "no hubo error": una fila que
        perdiera los tres en el camino también daría cero errores."""
        f = prev_mod.preview(_archivo_viejo()).filas_validas[0]

        assert f.tipo == "operativo", "el default de PRODUCTO, no el 'anual' de la columna"
        assert f.periodicidad == ""
        assert f.areas_involucradas == []

    def test_y_llegan_asi_al_payload_de_alta(self, users) -> None:
        """El último punto antes de la base: lo que `_a_create` le pasa a `ObjetivoService`."""
        f = prev_mod.preview(_archivo_viejo()).filas_validas[0]

        create = ObjetivosImportService._a_create(f, UUID(EMPRESA))

        assert (create.tipo, create.periodicidad, create.areas_involucradas) == (
            "operativo", "", [])


# ── 2. 🔴 La columna Tipo ─────────────────────────────────────────────────────

class TestLaColumnaTipo:

    def _por_titulo(self, users) -> dict:
        return {f.titulo: f for f in prev_mod.preview(_archivo_nuevo()).filas_validas}

    def test_un_tipo_VALIDO_se_respeta(self, users) -> None:
        """🔴 EL CONTRAPESO del test de abajo, y va primero: sin él, un `parsear_fila` que
        ignorara la columna y devolviera siempre 'operativo' pasaría el test del inválido."""
        assert self._por_titulo(users)["Plan anual"].tipo == "anual"

    def test_un_tipo_con_MAYUSCULA_entra(self, users) -> None:
        """"Anual" es lo que escribe una persona. El `.lower()` lo resuelve; sin él caería al
        default y el usuario vería su objetivo en la otra vista."""
        assert self._por_titulo(users)["Revisar convenio"].tipo == "anual"

    def test_un_tipo_INVALIDO_cae_al_default_y_la_fila_ENTRA(self, users) -> None:
        """🔴 EL TEST DE LA SESIÓN. Es la simetría con `Prioridad`: rechazar la fila entera por
        una celda mal escrita sería desproporcionado — el dato importante es el título y el
        responsable. Las DOS mitades importan: el valor cae Y la fila no se pierde."""
        res = prev_mod.preview(_archivo_nuevo())

        fila = {f.titulo: f for f in res.filas_validas}.get("Cerrar el mes")
        assert fila is not None, "la fila se rechazó en vez de caer al default"
        assert fila.tipo == "operativo"
        assert not [e for e in res.errores if e.identificador == "Cerrar el mes"]

    def test_la_periodicidad_viaja_tal_cual(self, users) -> None:
        """Texto libre: no se normaliza, no se valida contra nada."""
        assert self._por_titulo(users)["Cerrar el mes"].periodicidad == "mensual"

    def test_el_tipo_del_Excel_llega_al_payload_de_alta(self, users) -> None:
        """🔴 Si `_a_create` omitiera `tipo`, el default lo pondría el schema y TODO nacería
        operativo: el preview mostraría "anual" y la base guardaría otra cosa. La pantalla
        confirmando lo que no pasó es el peor modo de falla de un import."""
        f = self._por_titulo(users)["Plan anual"]

        assert ObjetivosImportService._a_create(f, UUID(EMPRESA)).tipo == "anual"


# ── 3. 🔴 La columna de áreas y su separador ──────────────────────────────────

class TestElSeparadorDeAreas:

    @pytest.mark.parametrize("celda,esperado", [
        ("", []),
        ("Sistemas", ["Sistemas"]),
        ("Sistemas; Legales", ["Sistemas", "Legales"]),
        # 🔴 con `;` presente, la coma es CONTENIDO
        ("Legales, Compliance; Sistemas", ["Legales, Compliance", "Sistemas"]),
        # sin `;`, la coma separa (el caso copiar-y-pegar)
        ("Sistemas, Legales", ["Sistemas", "Legales"]),
        # el `;` colgado es la salida para un área con coma escrita sola
        ("Legales, Compliance;", ["Legales, Compliance"]),
        ("Sistemas; ; Legales", ["Sistemas", "Legales"]),   # no mete elementos vacíos
    ])
    def test_la_regla_de_prioridad(self, celda, esperado) -> None:
        assert separar_areas(celda) == esperado

    def test_LA_LIMITACION_DECLARADA_un_area_con_coma_y_sin_punto_y_coma_se_parte(self) -> None:
        """⚠️ Fijado a propósito para que nadie lo descubra como bug. Una celda con UNA sola área
        con coma y sin ningún `;` no tiene información para decidir, y entre las dos lecturas la
        más probable es que sean dos áreas. La salida está documentada: agregarle el `;`.

        Si este test empieza a fallar es porque alguien cambió la regla — y ahí hay que
        actualizar el encabezado de `_objetivos_import_valores`, no borrar el test."""
        assert separar_areas("Legales, Compliance") == ["Legales", "Compliance"]

    def test_un_area_con_coma_adentro_se_importa_como_UN_elemento(self, users) -> None:
        """🔴 EL TEST DE LA SESIÓN, de punta a punta: del .xlsx real al payload de alta."""
        f = {x.titulo: x for x in prev_mod.preview(_archivo_nuevo()).filas_validas}["Plan anual"]

        assert f.areas_involucradas == ["Legales, Compliance", "Sistemas"]
        create = ObjetivosImportService._a_create(f, UUID(EMPRESA))
        assert create.areas_involucradas == ["Legales, Compliance", "Sistemas"]
        assert "Compliance" not in create.areas_involucradas, "el área se partió en dos"

    def test_la_celda_con_COMA_como_separador_tambien_entra(self, users) -> None:
        """Contrapeso: una regla que sólo partiera por `;` dejaría esta celda como un solo área."""
        f = {x.titulo: x
             for x in prev_mod.preview(_archivo_nuevo()).filas_validas}["Revisar convenio"]

        assert f.areas_involucradas == ["Sistemas", "Legales"]

    def test_lo_que_EXPORTA_el_sistema_vuelve_a_entrar_IGUAL(self) -> None:
        """🔴 La ida y vuelta completa, y la razón por la que el export junta con `"; "`.
        RRHH exporta, edita en Excel y resube: si el import partiera por coma, el área
        "Legales, Compliance" que el export escribió entera volvería como dos."""
        original = ["Legales, Compliance", "Sistemas"]
        row = ObjetivoResponse(
            id="o-1", empresa_id=EMPRESA, responsable_id=ANA, titulo="X", prioridad="media",
            estado="por_hacer", created_at=datetime(2026, 1, 1), updated_at=datetime(2026, 1, 1),
            tipo="anual", areas_involucradas=original)

        celda = construir_filas_export([row])[0]["Áreas involucradas"]

        assert separar_areas(celda) == original


def _fila_como_la_arma_el_lector(export: dict) -> dict:
    """Una fila del export, con las claves normalizadas igual que `_import_excel.filas`.

    Es la forma exacta en la que llega una planilla exportada y resubida: el lector keyea por
    `normalizar_header`, que **conserva los acentos**.
    """
    from services._import_csv import normalizar_header
    return {normalizar_header(k): v for k, v in export.items()}


class TestLaPlanillaExportadaSeVuelveAReconocer:
    """🔴 ESTE BLOQUE NACIÓ DE UN BUG QUE ESCRIBÍ Y QUE EL TEST ENCONTRÓ.

    La primera versión afirmaba que `normalizar_header` iguala "Áreas involucradas" con
    "Areas involucradas". **No lo hace**: recorta, colapsa espacios y hace casefold, y los
    acentos los deja. O sea que la columna que el export escribe NO era la que el import buscaba,
    y una planilla exportada y resubida habría entrado sin áreas, en silencio y sin un error.
    Lo resuelve el reintento sin acentos de `_get`. Estos tests fijan ese contrato desde el lado
    que importa: **la fila que produce el export, leída por el import**.
    """

    def _export(self) -> dict:
        row = ObjetivoResponse(
            id="o-1", empresa_id=EMPRESA, responsable_id=ANA, titulo="X", prioridad="alta",
            estado="por_hacer", created_at=datetime(2026, 1, 1),
            updated_at=datetime(2026, 1, 1), tipo="anual", periodicidad="mensual",
            areas_involucradas=["Legales, Compliance", "Sistemas"])
        return construir_filas_export([row])[0]

    def test_el_export_escribe_las_areas_CON_acento(self) -> None:
        """La guarda del bloque: si el export dejara de acentuar, los tests de abajo pasarían
        por el camino fácil y no probarían el reintento."""
        assert "Áreas involucradas" in self._export()

    def test_normalizar_header_NO_iguala_las_dos_grafias(self) -> None:
        """El hecho que hace falta el reintento, fijado explícitamente. Si algún día
        `normalizar_header` empieza a sacar acentos, este test rojea y avisa que el reintento de
        `_get` pasó a ser código muerto."""
        from services._import_csv import normalizar_header
        from services._objetivos_import_transforms import COL_AREAS

        assert normalizar_header("Áreas involucradas") != normalizar_header(COL_AREAS)

    def test_el_import_las_encuentra_igual(self) -> None:
        """🔴 Lo que importa: los tres campos sobreviven la vuelta completa export → import."""
        f = parsear_fila(_fila_como_la_arma_el_lector(self._export()))

        assert f["areas"] == ["Legales, Compliance", "Sistemas"]
        assert f["tipo"] == "anual"
        assert f["periodicidad"] == "mensual"

    def test_tambien_encuentra_la_DESCRIPCION_acentuada(self) -> None:
        """El reintento no es sólo para las áreas: "Descripción" tenía el mismo problema y era
        anterior a esta sesión."""
        export = {**self._export(), "Descripción": "Un texto"}

        assert parsear_fila(_fila_como_la_arma_el_lector(export))["descripcion"] == "Un texto"

    def test_una_columna_que_NO_esta_sigue_dando_vacio(self) -> None:
        """Contrapeso: un `_get` que devolviera la primera clave parecida pasaría todo lo de
        arriba y rompería cualquier columna ausente."""
        assert parsear_fila({"titulo": "X"})["areas"] == []

    def test_TITULO_sigue_rebotando_el_archivo_entero(self) -> None:
        """🚩 LA LIMITACIÓN QUE ESTA SESIÓN **NO** ARREGLA, fijada para que no se descubra sola.

        `Titulo` es columna REQUERIDA y su chequeo lo hace `_import_excel.faltantes`, que compara
        con `normalizar_header` a secas — antes de que `_get` entre en juego. El export escribe
        "Título", así que **resubir una planilla exportada rebota entera** con COLUMNAS_FALTANTES.
        Es anterior a esta sesión y arreglarlo toca el chequeo compartido de columnas requeridas
        de los tres imports del repo.

        Si este test empieza a fallar es porque alguien lo arregló: sacarlo entonces, y borrar la
        nota de `_get`."""
        from services._import_excel import faltantes
        from services._objetivos_import_transforms import REQUERIDAS

        headers = list(self._export().keys())

        assert "Titulo" in faltantes(headers, REQUERIDAS)


# ── 4. 🔴 El desplegable de áreas ya usadas ───────────────────────────────────

# Punto 4: dos filas COMPARTEN "Sistemas", una tiene el array vacío, y hay un área con coma para
# verificar que el pool la ofrece entera. La segunda empresa prueba el acotado.
_OTRA_EMPRESA = str(uuid4())
_FILAS_POOL = [
    {"empresa_id": EMPRESA, "areas_involucradas": ["Sistemas", "Legales, Compliance"]},
    {"empresa_id": EMPRESA, "areas_involucradas": ["Sistemas", "administracion"]},
    {"empresa_id": EMPRESA, "areas_involucradas": []},
    {"empresa_id": _OTRA_EMPRESA, "areas_involucradas": ["Comercial"]},
]


@pytest.fixture
def pool(monkeypatch):
    import repositories.objetivo_areas_repo as mod
    fake = FakeSupabase({"objetivos": _FILAS_POOL})
    monkeypatch.setattr(mod, "supabase_admin", fake)
    return ObjetivoCatalogosService(repo=ObjetivoAreasRepo()), fake


def test_el_catalogo_del_pool_repite_un_area_y_tiene_una_fila_vacia() -> None:
    """Punto 4. Sin la repetición no se puede ver que deduplica; sin la vacía, que no aporta ''."""
    todas = [a for f in _FILAS_POOL for a in f["areas_involucradas"]]
    assert todas.count("Sistemas") == 2
    assert any(f["areas_involucradas"] == [] for f in _FILAS_POOL)


class TestElDesplegableDeAreas:

    def test_ofrece_areas_SUELTAS_y_no_la_celda_entera(self, pool) -> None:
        """🔴 EL TEST DE LA SESIÓN, y la ganancia entera de haber pasado la columna a `text[]`.
        Con la columna en texto, el pool habría ofrecido "Sistemas; Legales, Compliance" como UNA
        opción: un desplegable de COMBINACIONES, donde filtrar por "Legales, Compliance" a secas
        no habría estado disponible nunca."""
        svc, _ = pool

        areas = svc.get_areas_conocidas(EMPRESA)

        assert areas == ["administracion", "Legales, Compliance", "Sistemas"]
        assert not any(";" in a for a in areas), "el pool ofrece una celda sin aplanar"

    def test_no_repite_un_area_usada_por_dos_objetivos(self, pool) -> None:
        svc, _ = pool

        assert svc.get_areas_conocidas(EMPRESA).count("Sistemas") == 1

    def test_una_fila_SIN_areas_no_aporta_nada(self, pool) -> None:
        """Ni un `''` ni un `None` en el desplegable."""
        svc, _ = pool

        assert all(a.strip() for a in svc.get_areas_conocidas(EMPRESA))

    def test_ordena_sin_distinguir_MAYUSCULAS(self, pool) -> None:
        """Con el orden de bytes, "administracion" caería después de "Sistemas" y el desplegable
        se leería al azar."""
        svc, _ = pool

        areas = svc.get_areas_conocidas(EMPRESA)

        assert areas == sorted(areas, key=str.casefold)
        assert areas != sorted(areas), "el orden por bytes daría otra cosa: el test no discrimina"

    def test_acota_por_EMPRESA(self, pool) -> None:
        """Un filtro que ofrece opciones que no pueden dar resultado en esta vista devuelve vacío
        sin explicar por qué."""
        svc, _ = pool

        assert "Comercial" not in svc.get_areas_conocidas(EMPRESA)

    def test_en_CONSOLIDADO_trae_las_de_todas_las_empresas(self, pool) -> None:
        """Contrapeso del anterior: un repo que filtrara SIEMPRE lo pasaría igual."""
        svc, _ = pool

        assert "Comercial" in svc.get_areas_conocidas(None)

    def test_la_empresa_viaja_en_la_QUERY_y_no_se_filtra_en_python(self, pool) -> None:
        """Forma A del patrón de barrera: filtrar después de traer las filas significa traerse
        los objetivos de todas las empresas por la red en cada apertura del desplegable."""
        svc, fake = pool

        svc.get_areas_conocidas(EMPRESA)

        assert ("objetivos", "empresa_id", [EMPRESA]) in fake.consultas


# ── 5. 🔴 El duplicado, con mensaje legible ───────────────────────────────────

class _APIErrorFalsa(Exception):
    """Como la `APIError` de postgrest: `.code` con el SQLSTATE y el texto crudo en el mensaje."""

    def __init__(self) -> None:
        self.message = ('duplicate key value violates unique constraint '
                        '"ux_objetivo_responsable_titulo"')
        self.code = "23505"
        super().__init__(self.message)


class _RepoQueRebotaLaSegunda:
    """Acepta la primera fila y choca el índice único en la segunda.

    🔴 LAS DOS COSAS IMPORTAN. Si rebotara siempre, no se podría verificar que el lote NO aborta;
    si no rebotara nunca, la rama del duplicado sería inalcanzable. Es el molde del
    `_ObjetivosFake` de `test_objetivos_import`, con el choque real en vez de un `AppError`.
    """

    def __init__(self) -> None:
        self.altas: list = []

    def find_by_id(self, id, empresa_id=None):
        return None

    def tiene_hijos(self, id, empresa_id=None) -> bool:
        return False

    def save(self, data):
        if len(self.altas) >= 1:
            raise _APIErrorFalsa()
        self.altas.append(data)
        return ObjetivoResponse(
            id=str(uuid4()), empresa_id=str(data.empresa_id),
            responsable_id=str(data.responsable_id), titulo=data.titulo,
            prioridad=data.prioridad, estado="por_hacer",
            created_at=datetime(2026, 1, 1), updated_at=datetime(2026, 1, 1),
            tipo=data.tipo, periodicidad=data.periodicidad,
            areas_involucradas=list(data.areas_involucradas))


@pytest.fixture
def import_con_duplicado(monkeypatch):
    """El `ObjetivoService` REAL sobre un repo que rebota: así corre la traducción de verdad."""
    import services.objetivo_service as svc_mod
    monkeypatch.setattr(svc_mod, "ensure_responsable_valido", lambda _id: None)
    repo = _RepoQueRebotaLaSegunda()
    return ObjetivosImportService(objetivos=ObjetivoService(repo=repo),
                                  audit=SimpleNamespace(registrar=lambda **k: None)), repo


def _body(users_fixture) -> ImportacionObjetivosConfirmarRequest:
    filas = prev_mod.preview(_archivo_viejo()).filas_validas
    return ImportacionObjetivosConfirmarRequest(empresa_id=UUID(EMPRESA), filas=filas)


class TestElDuplicadoSeReportaLegible:

    def test_el_motivo_dice_QUE_HACER(self, users, import_con_duplicado) -> None:
        """🔴 EL TEST DE LA SESIÓN. Con la clave de cinco columnas, el mensaje puede nombrar las
        tres salidas: cambiar el título, la periodicidad, o la vista."""
        svc, _ = import_con_duplicado

        res = svc.confirmar(_body(users))

        assert len(res.errores) == 1
        motivo = res.errores[0].motivo
        for palabra in ("título", "periodicidad", "vista"):
            assert palabra in motivo, f"el motivo no nombra «{palabra}»: {motivo!r}"

    def test_el_motivo_NO_contiene_el_texto_de_POSTGRES(self, users,
                                                        import_con_duplicado) -> None:
        """🔴 La aserción negativa, y la que separa esto de "hay algún mensaje". Hasta la sesión
        1 el motivo era literalmente `duplicate key value violates unique constraint
        "ux_objetivo_responsable_titulo"`, que no le dice nada a alguien de RRHH y encima expone
        el nombre de un objeto de la base."""
        svc, _ = import_con_duplicado

        motivo = svc.confirmar(_body(users)).errores[0].motivo

        assert "duplicate key" not in motivo
        assert "ux_objetivo_responsable_titulo" not in motivo
        assert "23505" not in motivo

    def test_el_lote_NO_aborta_por_la_fila_duplicada(self, users, import_con_duplicado) -> None:
        """La otra fila entra igual. Es la regla del flujo: el lote no aborta por una fila."""
        svc, repo = import_con_duplicado

        res = svc.confirmar(_body(users))

        assert res.importados == 1
        assert len(repo.altas) == 1

    def test_la_fila_se_identifica_por_su_TITULO(self, users, import_con_duplicado) -> None:
        """Para encontrarla en la planilla hace falta el título, no el número de fila solo."""
        svc, _ = import_con_duplicado

        assert svc.confirmar(_body(users)).errores[0].identificador == "Auditar licencias"

    def test_un_error_que_NO_es_duplicado_sigue_subiendo(self, users, monkeypatch) -> None:
        """Contrapeso: si `duplicado_a_409` tradujera cualquier excepción, este también saldría
        como "ya existe un objetivo con ese título" — y sería mentira."""
        import services.objetivo_service as svc_mod
        monkeypatch.setattr(svc_mod, "ensure_responsable_valido", lambda _id: None)

        class _RepoRoto(_RepoQueRebotaLaSegunda):
            def save(self, data):
                raise RuntimeError("connection timeout")

        svc = ObjetivosImportService(objetivos=ObjetivoService(repo=_RepoRoto()),
                                     audit=SimpleNamespace(registrar=lambda **k: None))

        motivos = [e.motivo for e in svc.confirmar(_body(users)).errores]

        assert all("timeout" in m for m in motivos)


def test_parsear_fila_no_rompe_con_una_fila_completamente_vacia() -> None:
    """Borde: `_get` devuelve '' para todo y los tres campos nuevos caen a sus defaults sin que
    ninguno reviente. Vale la pena porque `separar_areas` y el `.lower()` del tipo son los dos
    primeros que tocarían un `None`."""
    f = parsear_fila({})

    assert (f["tipo"], f["periodicidad"], f["areas"]) == ("operativo", "", [])
