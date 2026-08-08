"""
LECTOR de un Excel subido: hoja, cabeceras y numeración de filas. Simétrico a `_import_csv`.

Devuelve **filas como dicts con los headers como claves y valores string**, exactamente la misma
forma que `_import_csv.filas`. Eso es deliberado: los parsers de cada flujo no tienen que saber
de qué formato vino el archivo, y el día que un import acepte los dos solo cambia el lector.

🔴 NO HAY DETECCIÓN DE ENCODING ACÁ, Y NO ES UN OLVIDO. Un `.xlsx` es un ZIP con XML adentro y
el XML declara su propio encoding: `_import_encoding` —con sus cuatro pasos, su heurística de
UTF-16 y su `permitir_latin1`— **no aplica y no se importa**. Todo ese módulo existe para
resolver un problema que el formato binario ya tiene resuelto.

## 🔴 LAS SEIS TRAMPAS DE EXCEL QUE EL CSV NO TIENE

En un CSV todo es texto: lo que se lee es lo que el usuario escribió. En un Excel no, y cada
una de estas convierte un archivo correcto en datos rotos SIN error:

  1. **Celda vacía = `None`, no `""`.** `None.strip()` revienta y `str(None)` da `"None"` — que
     entraría a la base como el texto "None". Se normaliza a `""` para igualar al CSV.
  2. **Números.** Excel guarda todo como numérico: un DNI escrito `12345678` puede volver como
     float y `str()` lo convierte en `"12345678.0"`. Un dedup por DNI contra ese string no
     matchea NUNCA y el import crea duplicados en silencio. Los integrales se emiten sin `.0`.
  3. **Fechas = `datetime`, no string.** Y `datetime`, no `date`, aun cuando la celda muestre
     solo el día (medido con openpyxl 3.1.5). Se emiten como `d/m/Y`, que es el formato que ya
     saben leer los parsers del repo (`_nomina_parsers.parse_fecha`).
  4. **Espacios invisibles en los headers.** `"Titulo "` con espacio final es indistinguible de
     `"Titulo"` en pantalla. Se normaliza con la MISMA función que el CSV
     (`_import_csv.normalizar_header`) para que las dos rutas acepten y rechacen lo mismo.
  5. **Varias hojas.** Se lee **SOLO la primera** (`wb.worksheets[0]`), no la "activa": la
     activa es la que estaba abierta cuando se guardó, o sea que el archivo se leería distinto
     según en qué pestaña quedó el cursor de quien lo mandó. La primera es estable y es lo que
     el usuario ve al abrirlo.
  6. **Filas vacías al final.** Excel infla `max_row` con filas que se tocaron y se borraron, y
     `iter_rows` las devuelve enteras en `None`. Se descartan: una fila sin ningún valor no es
     un dato, y sin este filtro el reporte diría "50 filas con error" sobre un archivo de 3.

⚠️ **`data_only=True` devuelve el ÚLTIMO VALOR CALCULADO, no la fórmula — y si el productor
nunca calculó, devuelve `None`.** Medido: un `.xlsx` generado por una librería (no por Excel) no
trae valor cacheado y la celda con fórmula llega vacía. Es el comportamiento correcto para
nuestro caso —queremos el valor, no `=SUMA(A1:A3)`— pero hay que saberlo: un archivo armado por
un script con fórmulas se lee como columnas vacías, no como error.
"""
import io
from datetime import date, datetime
from typing import Iterator, List, Tuple

from openpyxl import load_workbook

from services._import_csv import normalizar_header


def _valor(v) -> str:
    """Celda → string, resolviendo las trampas 1, 2 y 3 del encabezado del módulo."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"      # antes que int: en Python bool ES int
    if isinstance(v, float) and v.is_integer():
        return str(int(v))                   # 12345678.0 → "12345678"
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")        # el formato que ya leen los parsers del repo
    return str(v).strip()


def abrir(data: bytes) -> Tuple[List[str], List[list]]:
    """Devuelve `(headers, filas_crudas)` de la PRIMERA hoja. Los headers salen sin normalizar.

    Se devuelven los headers tal cual vinieron porque el mensaje de error los muestra a alguien
    que tiene la planilla abierta; la comparación se hace normalizada más abajo.

    Raises:
        ValueError: si el archivo no es un Excel legible o no tiene ninguna hoja.
    """
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl levanta de todo ante un archivo corrupto
        raise ValueError(
            "No se pudo leer el archivo. ¿Es un Excel (.xlsx) válido? "
            "Si lo exportaste de otro sistema, abrilo y guardalo desde Excel."
        ) from exc
    if not wb.worksheets:
        raise ValueError("El archivo no tiene ninguna hoja.")
    ws = wb.worksheets[0]                     # la PRIMERA, no la activa: ver la trampa 5
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return [], []
    headers = [_valor(c) for c in filas[0]]
    return headers, filas[1:]


def faltantes(headers: List[str], requeridas: List[str]) -> List[str]:
    """Las columnas requeridas que NO están, con su nombre ORIGINAL para el mensaje.

    Misma semántica y misma normalización que `_import_csv.faltantes`: las dos rutas tienen que
    aceptar y rechazar exactamente los mismos encabezados.
    """
    presentes = {normalizar_header(h) for h in headers if h}
    return [c for c in requeridas if normalizar_header(c) not in presentes]


def filas(headers: List[str], crudas: List[list]) -> Iterator[Tuple[int, dict]]:
    """Rinde `(nº de fila del archivo, fila)` con claves normalizadas y valores string.

    El número arranca en 2 porque la fila 1 es el encabezado: es el número que ve quien abre el
    Excel, y por eso el reporte de errores es accionable — igual que en `_import_csv.filas`.

    Las filas completamente vacías se DESCARTAN (trampa 6) pero **no rompen la numeración**: si
    la fila 7 está vacía, la 8 sigue siendo la 8 para quien mira la planilla.
    """
    claves = [normalizar_header(h) for h in headers]
    for n, cruda in enumerate(crudas, start=2):
        valores = [_valor(c) for c in cruda]
        if not any(valores):
            continue                          # fila fantasma de Excel, no un dato
        yield n, {k: v for k, v in zip(claves, valores) if k}


def hojas(data: bytes) -> List[str]:
    """Nombres de las hojas del archivo. Para poder decirle al usuario cuál se leyó."""
    try:
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True).sheetnames
    except Exception:  # noqa: BLE001
        return []
